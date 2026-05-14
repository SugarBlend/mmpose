from __future__ import annotations
from math import pi
from pathlib import Path
import math
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import openpyxl.cell.cell
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from mmpose.utils.logger import MMLogger


matplotlib.use("tkAgg")
logger = MMLogger.get_instance("Accuracy")

STATS_NAMES = ["AP", "AP .5", "AP .75", "AP (M)", "AP (L)",
               "AR", "AR .5", "AR .75", "AR (M)", "AR (L)"]

WHOLEBODY_SECTIONS: dict[str, str] = {
    "body": "keypoints_body",
    "foot": "keypoints_foot",
    "face": "keypoints_face",
    "left_hand": "keypoints_lefthand",
    "right_hand": "keypoints_righthand",
    "all": "keypoints_wholebody",
}

SECTION_DISPLAY_LABELS: dict[str, str] = {
    "body": "Body (17 kps)",
    "foot": "Foot (6 kps)",
    "face": "Face (68 kps)",
    "left_hand": "Left Hand (21 kps)",
    "right_hand": "Right Hand (21 kps)",
    "all": "Wholebody (133 kps)",
}


def _split_wholebody_metrics(raw: dict) -> dict[str, dict[str, float]]:
    sections: dict[str, dict[str, float]] = {}
    for section_key, metrics_key in WHOLEBODY_SECTIONS.items():
        if metrics_key in raw and isinstance(raw[metrics_key], dict):
            sections[section_key] = dict(raw[metrics_key])
    return sections


def save_metrics_xlsx(
    exp_metrics: dict[str, dict[str, float]],
    save_path: str,
    is_wholebody: bool = True,
) -> None:
    wb = Workbook()
    default_sheet = wb.active

    header_font = Font(name="TimesNewRoman", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2F5496")
    subhdr_fill = PatternFill("solid", start_color="D6E4F7")
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_header(cell: openpyxl.cell.cell.Cell, main: bool = True) -> None:
        cell.font = header_font if main else Font(name="TimesNewRoman", bold=True)
        cell.fill = header_fill if main else subhdr_fill
        cell.alignment = center_align
        cell.border = thin_border

    def _write_section_sheet(
        sheet_name: str,
        section_data: dict[str, dict[str, float]],
    ) -> None:
        ws = wb.create_sheet(sheet_name)
        legends = list(section_data.keys())

        _style_header(ws.cell(1, 1, "Metric"))
        for col_idx, legend in enumerate(legends, start=2):
            _style_header(ws.cell(1, col_idx, legend))

        for row_idx, stat in enumerate(STATS_NAMES, start=2):
            cell = ws.cell(row_idx, 1, stat)
            cell.font = Font(name="TimesNewRoman", bold=True)
            cell.fill = subhdr_fill
            cell.border = thin_border

            for col_idx, legend in enumerate(legends, start=2):
                val = section_data[legend].get(stat)
                cell = ws.cell(row_idx, col_idx, round(val, 4) if val else "—")
                cell.alignment = center_align
                cell.border = thin_border
                cell.font = Font(name="TimesNewRoman")

        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, len(legends) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    if is_wholebody:
        by_section: dict[str, dict[str, dict[str, float]]] = {
            s: {} for s in WHOLEBODY_SECTIONS
        }
        for legend, raw in exp_metrics.items():
            split = _split_wholebody_metrics(raw)
            for section, stats_dict in split.items():
                by_section[section][legend] = stats_dict

        for section, label in SECTION_DISPLAY_LABELS.items():
            if by_section.get(section):
                _write_section_sheet(label, by_section[section])
    else:
        _write_section_sheet("Metrics", exp_metrics)

    wb.remove(default_sheet)
    wb.save(save_path)
    logger.info(f"Metrics XLSX saved: '{save_path}'")


def _draw_radar(ax, labels, datasets, xticks, colors):
    N = len(labels)
    angles = [n / N * 2 * pi for n in range(N)] + [0]

    ax.set_theta_offset(pi / 2)
    ax.set_theta_direction(-1)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, size=7)

    ax.set_yticks(xticks)
    ax.set_yticklabels([f"{t:.2f}" for t in xticks], size=6)

    margin = (xticks[-1] - xticks[0]) * 0.08
    ax.set_ylim(min(xticks), max(xticks) + margin)
    ax.yaxis.set_major_locator(matplotlib.ticker.FixedLocator(xticks))

    for idx, (name, values) in enumerate(datasets):
        vals = values + [values[0]]
        color = colors[idx % len(colors)]
        ax.plot(angles, vals, linewidth=1.8, linestyle="solid", label=name, color=color)
        ax.fill(angles, vals, color=color, alpha=0.12)


def _auto_xticks(values_flat: list[float], n_steps: int = 5) -> list[float]:
    if not values_flat:
        return [0.0, 0.2, 0.4, 0.6, 0.8, 1.0]

    lo = min(v for v in values_flat if v > 0)
    hi = max(values_flat)

    span = hi - lo
    pad  = max(0.02, span * 0.1)

    lo = max(0.0, lo - pad)
    hi = hi + pad

    # Round to a multiple of 0.05
    lo = int(lo * 20) / 20
    hi = math.ceil(hi * 20) / 20

    # step multiple of 0.05
    raw_step = (hi - lo) / n_steps
    step = max(0.05, round(raw_step * 20) / 20)

    # Recalculate taking into account the rounded step
    ticks = [round(lo + step * i, 2) for i in range(n_steps + 1)]

    # If the maximum data is still higher than the last tick, we add another one.
    while ticks[-1] < hi:
        ticks.append(round(ticks[-1] + step, 2))

    return ticks


def generate_radar_plot(
    exp_metrics: dict[str, dict[str, float]],
    is_wholebody: bool = True,
    xticks: list[float] | None = None,
    save_path: str | None = None,
    show: bool = True,
    title: str = "Pose Model Comparison",
) -> plt.Figure | None:
    if not exp_metrics:
        logger.warning("exp_metrics is empty — skip radar.")
        return None

    colors = plt.cm.tab10.colors

    if is_wholebody:
        by_section: dict[str, dict[str, dict[str, float]]] = {
            s: {} for s in WHOLEBODY_SECTIONS
        }
        for legend, raw in exp_metrics.items():
            split = _split_wholebody_metrics(raw)
            for section, stats_dict in split.items():
                if section in by_section:
                    by_section[section][legend] = stats_dict

        non_empty = [
            (s, lbl)
            for s, lbl in SECTION_DISPLAY_LABELS.items()
            if by_section.get(s)
        ]
        if not non_empty:
            logger.warning("Doesn't have enough data for wholebody radar.")
            return None

        ncols = 3
        nrows = (len(non_empty) + ncols - 1) // ncols
        fig, axes = plt.subplots(
            nrows, ncols,
            figsize=(7 * ncols, 7 * nrows),
            subplot_kw=dict(polar=True),
        )
        axes_flat = np.array(axes).flatten()

        for plot_idx, (section, section_label) in enumerate(non_empty):
            ax = axes_flat[plot_idx]
            sec_data = by_section[section]      # {legend: {stat: val}}
            datasets = [
                (legend, [sec_data[legend].get(s, 0.0) for s in STATS_NAMES])
                for legend in sec_data
            ]

            all_vals = [v for _, vals in datasets for v in vals if v > 0]
            data_max = max(all_vals)
            section_ticks = xticks if xticks is not None else _auto_xticks(all_vals)

            # Guarantee that the last tick is higher than the data maximum
            if section_ticks[-1] < data_max:
                step = section_ticks[-1] - section_ticks[-2]
                while section_ticks[-1] < data_max:
                    section_ticks.append(round(section_ticks[-1] + step, 2))

            _draw_radar(ax, STATS_NAMES, datasets, section_ticks, colors)
            ax.set_title(section_label, pad=2, size=11, weight="bold")
            ax.legend(loc="upper right", bbox_to_anchor=(1.35, 1.15), fontsize=7)

        # Hiding unnecessary subplots
        for extra_idx in range(len(non_empty), len(axes_flat)):
            axes_flat[extra_idx].set_visible(False)

        fig.suptitle(title, size=15, weight="bold", y=1.)

    else:
        # Single radar for non-wholebody model
        all_stats = list(next(iter(exp_metrics.values())).keys())
        if not all_stats:
            logger.warning("Metrics is empty - skip radar creation.")
            return None

        fig, ax = plt.subplots(figsize=(8, 8), subplot_kw=dict(polar=True))
        datasets = [
            (legend, [m.get(s, 0.0) for s in all_stats])
            for legend, m in exp_metrics.items()
        ]

        all_vals = [v for _, vals in datasets for v in vals if v > 0]
        data_max = max(all_vals)
        section_ticks = xticks if xticks is not None else _auto_xticks(all_vals)

        # Guarantee that the last tick is higher than the data maximum
        if section_ticks[-1] < data_max:
            step = section_ticks[-1] - section_ticks[-2]
            while section_ticks[-1] < data_max:
                section_ticks.append(round(section_ticks[-1] + step, 2))

        _draw_radar(ax, STATS_NAMES, datasets, section_ticks, colors)
        ax.set_title(title, pad=2, size=13, weight="bold")
        ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1))

    fig.tight_layout()
    # fig.tight_layout(rect=[0, 0, 1, 0.97])

    if save_path:
        Path(save_path).parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(save_path, dpi=200, bbox_inches="tight")
        logger.info(f"Plot saved → {save_path}")

    if show:
        plt.show(block=True)

    return fig
